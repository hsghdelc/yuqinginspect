import copy
import threading
import tkinter as tk
from pathlib import Path
from tkinter import filedialog, messagebox, scrolledtext, ttk

try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
except Exception:
    DND_FILES = None
    TkinterDnD = None

from monthly_special import get_default_monthly_plans
from processor import DEFAULT_SCHEME, DEFAULT_SCHEME_ID, copy_json, load_config, process_file, save_config


BaseTk = TkinterDnD.Tk if TkinterDnD else tk.Tk


class ReviewTool(BaseTk):
    def __init__(self):
        super().__init__()
        self.title("南方分中心舆情质检辅助工具")
        self.geometry("860x620")
        self.minsize(820, 560)

        self.input_path = tk.StringVar()
        self.output_dir = tk.StringVar(value=str(Path.home() / "Desktop"))
        self.inspector = tk.StringVar()
        self.status = tk.StringVar(value="请选择或拖入舆情质检明细 Excel 文件")
        self.scheme_var = tk.StringVar()
        self.scheme_options = []
        self.last_report_text = ""
        self.copy_button = None

        self._build_ui()
        self.refresh_schemes()
        self._enable_drop()

    def _build_ui(self):
        self.columnconfigure(0, weight=1)
        self.rowconfigure(2, weight=1)

        header = ttk.Frame(self, padding=(18, 16, 18, 8))
        header.grid(row=0, column=0, sticky="ew")
        header.columnconfigure(0, weight=1)
        ttk.Label(
            header,
            text="南方分中心舆情质检辅助工具",
            font=("Microsoft YaHei UI", 15, "bold"),
        ).grid(row=0, column=0, sticky="w")
        ttk.Button(header, text="方案配置", command=self.open_scheme_config).grid(row=0, column=1, padx=(10, 0))
        ttk.Button(header, text="专项配置", command=self.open_special_config).grid(row=0, column=2, padx=(8, 0))

        form = ttk.LabelFrame(self, text="文件与人员", padding=(14, 12))
        form.grid(row=1, column=0, sticky="ew", padx=18, pady=(0, 12))
        form.columnconfigure(1, weight=1)
        form.columnconfigure(2, minsize=88)

        self._row(form, 0, "源文件", self.input_path, self.choose_file)
        self._row(form, 1, "输出目录", self.output_dir, self.choose_output_dir)
        ttk.Label(form, text="质检方案", width=10, anchor="e").grid(row=2, column=0, sticky="e", padx=(0, 10), pady=7)
        self.scheme_combo = ttk.Combobox(form, textvariable=self.scheme_var, state="readonly")
        self.scheme_combo.grid(row=2, column=1, sticky="ew", pady=7)

        ttk.Label(form, text="质检人员", width=10, anchor="e").grid(row=3, column=0, sticky="e", padx=(0, 10), pady=7)
        ttk.Entry(form, textvariable=self.inspector).grid(row=3, column=1, sticky="ew", pady=7)

        self.drop_tip = ttk.Label(
            form,
            text="可将 .xlsx / .xlsm 文件拖入窗口；如拖入不可用，请点击“选择”。",
            foreground="#546179",
        )
        self.drop_tip.grid(row=4, column=1, columnspan=2, sticky="w", pady=(4, 0))

        body = ttk.Frame(self, padding=(18, 0, 18, 0))
        body.grid(row=2, column=0, sticky="nsew")
        body.columnconfigure(0, weight=1)
        body.rowconfigure(1, weight=1)

        actions = ttk.Frame(body)
        actions.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        self.run_button = ttk.Button(actions, text="开始处理", command=self.run)
        self.run_button.pack(side="left")
        ttk.Button(actions, text="清空日志", command=lambda: self.log.delete("1.0", "end")).pack(side="left", padx=8)
        self.copy_button = ttk.Button(actions, text="复制日报送内容", command=self.copy_report, state="disabled")
        self.copy_button.pack(side="left")
        ttk.Label(actions, textvariable=self.status, foreground="#1f5fbf").pack(side="right")

        log_frame = ttk.LabelFrame(body, text="处理日志", padding=(8, 8))
        log_frame.grid(row=1, column=0, sticky="nsew")
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)
        self.log = scrolledtext.ScrolledText(log_frame, height=18, wrap="word")
        self.log.grid(row=0, column=0, sticky="nsew")
        self._log("使用说明：选择或拖入舆情质检明细 Excel，确认输出目录和质检人员后点击“开始处理”。处理完成后可一键复制日报送内容。")

    def _row(self, parent, row, label, var, command):
        ttk.Label(parent, text=label, width=10, anchor="e").grid(row=row, column=0, sticky="e", padx=(0, 10), pady=7)
        ttk.Entry(parent, textvariable=var).grid(row=row, column=1, sticky="ew", pady=7)
        ttk.Button(parent, text="选择", command=command).grid(row=row, column=2, sticky="ew", padx=(10, 0), pady=7)

    def _enable_drop(self):
        if not DND_FILES:
            self.drop_tip.config(text="当前运行环境未启用拖入文件；请点击“选择”添加源文件。")
            return
        self.drop_target_register(DND_FILES)
        self.dnd_bind("<<Drop>>", self._on_drop)

    def _on_drop(self, event):
        files = self.tk.splitlist(event.data)
        if not files:
            return
        path = Path(files[0])
        if path.suffix.lower() not in {".xlsx", ".xlsm"}:
            messagebox.showwarning("提示", "请拖入 .xlsx 或 .xlsm 文件")
            return
        self.input_path.set(str(path))
        self.status.set("已识别拖入文件")

    def choose_file(self):
        path = filedialog.askopenfilename(
            title="选择舆情质检明细",
            filetypes=[("Excel 文件", "*.xlsx *.xlsm"), ("所有文件", "*.*")],
        )
        if path:
            self.input_path.set(path)

    def choose_output_dir(self):
        path = filedialog.askdirectory(title="选择输出目录")
        if path:
            self.output_dir.set(path)

    def refresh_schemes(self):
        config = load_config()
        schemes = config.get("schemes") or {DEFAULT_SCHEME_ID: DEFAULT_SCHEME}
        active_id = config.get("active_scheme_id") or DEFAULT_SCHEME_ID
        self.scheme_options = [(scheme_id, scheme.get("name") or scheme_id) for scheme_id, scheme in schemes.items()]
        self.scheme_combo["values"] = [name for _scheme_id, name in self.scheme_options]
        selected_index = 0
        for idx, (scheme_id, _name) in enumerate(self.scheme_options):
            if scheme_id == active_id:
                selected_index = idx
                break
        if self.scheme_options:
            self.scheme_combo.current(selected_index)

    def selected_scheme_id(self):
        selected_name = self.scheme_var.get()
        for scheme_id, name in self.scheme_options:
            if name == selected_name:
                return scheme_id
        return DEFAULT_SCHEME_ID

    def open_scheme_config(self):
        SchemeConfigWindow(self)

    def open_special_config(self):
        SpecialConfigWindow(self)

    def run(self):
        input_path = self.input_path.get().strip()
        output_dir = self.output_dir.get().strip()
        if not input_path:
            messagebox.showwarning("提示", "请先选择或拖入源文件")
            return
        if not Path(input_path).exists():
            messagebox.showerror("错误", "源文件不存在")
            return
        if not output_dir:
            messagebox.showwarning("提示", "请选择输出目录")
            return

        self.last_report_text = ""
        self.copy_button.config(state="disabled")
        self.run_button.config(state="disabled")
        self.status.set("正在处理，请稍候...")
        self._log("开始处理：" + input_path)
        scheme_id = self.selected_scheme_id()
        inspector = self.inspector.get()
        threading.Thread(target=self._run_worker, args=(input_path, output_dir, scheme_id, inspector), daemon=True).start()

    def _run_worker(self, input_path, output_dir, scheme_id, inspector):
        try:
            result = process_file(
                input_path,
                output_dir,
                inspector,
                progress_callback=self._thread_log,
                scheme_id=scheme_id,
            )
            self.after(0, self._run_success, result)
        except Exception as exc:
            self.after(0, self._run_failed, exc)

    def _thread_log(self, text):
        self.after(0, self._log, text)

    def _run_success(self, result):
        self.run_button.config(state="normal")
        self.status.set("处理完成")
        self.last_report_text = result["report_text"]
        self.copy_button.config(state="normal")
        self._log("专项质检：" + result["special_name"])
        self._log(f"专项命中：{result['special_count']} 条")
        self._log(f"舆情提醒复核：{result['reminder_count']} 条")
        self._log(f"无效复核抽样：{result['invalid_count']} 条")
        if result["overtime_count"]:
            self._log("超时预警：" + str(result["overtime_count"]) + " 条，编号：" + "、".join(result["overtime_ids"]))
        else:
            self._log("超时预警：未查询到超时舆情")
        self._log("日报送文本：" + result["report_text"])
        self._log("输出文件：" + result["output_path"])
        messagebox.showinfo("完成", "舆情质检明细已生成。")

    def _run_failed(self, exc):
        self.run_button.config(state="normal")
        self.status.set("处理失败")
        self._log("处理失败：" + str(exc))
        messagebox.showerror("处理失败", str(exc))

    def copy_report(self):
        if not self.last_report_text:
            messagebox.showinfo("提示", "暂无可复制的日报送内容")
            return
        self.clipboard_clear()
        self.clipboard_append(self.last_report_text)
        self.status.set("日报送内容已复制")

    def _log(self, text):
        self.log.insert("end", text + "\n")
        self.log.see("end")


class SchemeConfigWindow(tk.Toplevel):
    FIELD_OPTIONS = [
        ("keep", "基础保留"),
        ("company", "公司筛选"),
        ("send_time", "派发时间"),
        ("process_time", "处理时间"),
        ("event_nature", "事件性质"),
        ("valid_scope", "舆情范围"),
        ("marketing", "营销类"),
        ("reminder", "舆情提醒"),
        ("duplicate", "重复事件"),
        ("event_category", "事件分类"),
        ("id", "编号"),
        ("special_targets", "专项关键词匹配"),
    ]

    VALUE_OPTIONS = [
        ("valid_yes", "符合舆情范围：是", "valid_scope", "是"),
        ("valid_no", "符合舆情范围：否", "valid_scope", "否"),
        ("marketing_yes", "营销类：是", "marketing", "是"),
        ("marketing_no", "营销类：否", "marketing", "否"),
        ("duplicate_yes", "重复事件：是", "duplicate", "是"),
        ("duplicate_no", "重复事件：否", "duplicate", "否"),
        ("negative_event", "负面事件值", "event_nature", "负面事件"),
        ("positive_event", "正面事件值", "event_nature", "正面事件"),
        ("reminder", "舆情提醒值", "reminder", "舆情提醒"),
        ("livelihood_event", "民生事件值", "event_category", "民生类舆情"),
    ]

    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent
        self.title("方案配置")
        self.geometry("900x650")
        self.minsize(840, 590)
        self.transient(parent)

        self.config_data = load_config()
        self.schemes = copy.deepcopy(self.config_data.get("schemes") or {DEFAULT_SCHEME_ID: DEFAULT_SCHEME})
        self.current_scheme_id = None
        self.header_choices = []
        self.name_var = tk.StringVar()
        self.sample_rate_var = tk.StringVar()
        self.sample_min_var = tk.StringVar()
        self.overtime_var = tk.StringVar()
        self.field_row_vars = []
        self.value_row_vars = []
        self.field_rows_frame = None
        self.value_rows_frame = None

        self._build_ui()
        self.refresh_scheme_list()

    def _build_ui(self):
        self.columnconfigure(1, weight=1)
        self.rowconfigure(0, weight=1)

        left = ttk.Frame(self, padding=(14, 14))
        left.grid(row=0, column=0, sticky="ns")
        ttk.Label(left, text="质检方案").pack(anchor="w")
        self.scheme_list = tk.Listbox(left, height=14, exportselection=False, width=20)
        self.scheme_list.pack(fill="y", expand=True, pady=(8, 10))
        self.scheme_list.bind("<<ListboxSelect>>", self.on_scheme_select)
        ttk.Button(left, text="新增方案", command=self.add_scheme).pack(fill="x", pady=(0, 6))
        ttk.Button(left, text="复制方案", command=self.copy_scheme).pack(fill="x", pady=(0, 6))
        ttk.Button(left, text="删除方案", command=self.delete_scheme).pack(fill="x")

        right = ttk.Frame(self, padding=(0, 14, 14, 14))
        right.grid(row=0, column=1, sticky="nsew")
        right.columnconfigure(0, weight=1)
        right.rowconfigure(1, weight=1)

        top = ttk.Frame(right)
        top.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        top.columnconfigure(1, weight=1)
        ttk.Label(top, text="方案名称").grid(row=0, column=0, sticky="e", padx=(0, 8))
        ttk.Entry(top, textvariable=self.name_var).grid(row=0, column=1, sticky="ew")
        ttk.Button(top, text="读取源文件表头", command=self.load_headers).grid(row=0, column=2, padx=(10, 0))

        notebook = ttk.Notebook(right)
        notebook.grid(row=1, column=0, sticky="nsew")

        fields_tab = ttk.Frame(notebook, padding=(12, 12))
        values_tab = ttk.Frame(notebook, padding=(12, 12))
        company_tab = ttk.Frame(notebook, padding=(12, 12))
        notebook.add(fields_tab, text="字段映射")
        notebook.add(values_tab, text="规则与抽样")
        notebook.add(company_tab, text="公司名单")

        fields_tab.columnconfigure(1, weight=1)
        fields_tab.rowconfigure(2, weight=1)
        field_actions = ttk.Frame(fields_tab)
        field_actions.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 8))
        ttk.Button(field_actions, text="新增字段", command=self.add_field_row).pack(side="left")
        ttk.Label(
            field_actions,
            text="用途决定程序逻辑，显示名称和列可按源表调整；专项关键词匹配列可填多个，如 C,K。",
            foreground="#546179",
        ).pack(side="left", padx=(10, 0))
        field_header = ttk.Frame(fields_tab)
        field_header.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(0, 4))
        field_header.columnconfigure(2, weight=1)
        field_header.columnconfigure(3, weight=2)
        ttk.Label(field_header, text="启用", width=4).grid(row=0, column=0, padx=(0, 6))
        ttk.Label(field_header, text="用途", width=22).grid(row=0, column=1, padx=(0, 6))
        ttk.Label(field_header, text="显示名称").grid(row=0, column=2, sticky="w", padx=(0, 6))
        ttk.Label(field_header, text="匹配列 / 字段列").grid(row=0, column=3, sticky="w", padx=(0, 6))
        self.field_rows_frame = self._make_scroll_area(fields_tab, row=2)

        values_tab.columnconfigure(0, weight=1)
        values_tab.rowconfigure(2, weight=1)
        value_actions = ttk.Frame(values_tab)
        value_actions.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        ttk.Button(value_actions, text="新增判断规则", command=self.add_value_row).pack(side="left")
        ttk.Label(
            value_actions,
            text="规则类型决定日报统计、无效抽样和超时筛查的判断口径；列和值均可维护。",
            foreground="#546179",
        ).pack(side="left", padx=(10, 0))
        value_header = ttk.Frame(values_tab)
        value_header.grid(row=1, column=0, sticky="ew", pady=(0, 4))
        value_header.columnconfigure(2, weight=1)
        value_header.columnconfigure(3, weight=2)
        value_header.columnconfigure(4, weight=1)
        ttk.Label(value_header, text="启用", width=4).grid(row=0, column=0, padx=(0, 6))
        ttk.Label(value_header, text="规则类型", width=24).grid(row=0, column=1, padx=(0, 6))
        ttk.Label(value_header, text="显示名称").grid(row=0, column=2, sticky="w", padx=(0, 6))
        ttk.Label(value_header, text="判断列").grid(row=0, column=3, sticky="w", padx=(0, 6))
        ttk.Label(value_header, text="匹配值").grid(row=0, column=4, sticky="w", padx=(0, 6))
        self.value_rows_frame = self._make_scroll_area(values_tab, row=2)

        sampling = ttk.LabelFrame(values_tab, text="无效复核抽样与超时筛查", padding=(10, 8))
        sampling.grid(row=3, column=0, sticky="ew", pady=(10, 0))
        sampling.columnconfigure(1, weight=1)
        sampling.columnconfigure(3, weight=1)
        sampling.columnconfigure(5, weight=1)
        ttk.Label(sampling, text="抽样比例").grid(row=0, column=0, sticky="e", padx=(0, 8))
        ttk.Entry(sampling, textvariable=self.sample_rate_var, width=10).grid(row=0, column=1, sticky="ew")
        ttk.Label(sampling, text="最少条数").grid(row=0, column=2, sticky="e", padx=(12, 8))
        ttk.Entry(sampling, textvariable=self.sample_min_var, width=10).grid(row=0, column=3, sticky="ew")
        ttk.Label(sampling, text="超时阈值(分钟)").grid(row=0, column=4, sticky="e", padx=(12, 8))
        ttk.Entry(sampling, textvariable=self.overtime_var, width=10).grid(row=0, column=5, sticky="ew")

        company_tab.columnconfigure(0, weight=1)
        company_tab.rowconfigure(0, weight=1)
        self.company_text = scrolledtext.ScrolledText(company_tab, height=18, wrap="word")
        self.company_text.grid(row=0, column=0, sticky="nsew")
        ttk.Label(company_tab, text="每行一个公司名称；公司筛选列的值在名单中才会保留。", foreground="#546179").grid(
            row=1, column=0, sticky="w", pady=(8, 0)
        )

        actions = ttk.Frame(right)
        actions.grid(row=2, column=0, sticky="ew", pady=(12, 0))
        ttk.Button(actions, text="保存当前方案", command=self.save_current_scheme).pack(side="left")
        ttk.Button(actions, text="设为默认方案", command=self.set_active_scheme).pack(side="left", padx=8)
        ttk.Button(actions, text="恢复当前方案默认值", command=self.reset_current_scheme).pack(side="left")
        ttk.Button(actions, text="保存全部并关闭", command=self.save_all_and_close).pack(side="right")

    def refresh_scheme_list(self):
        self.scheme_list.delete(0, "end")
        ids = list(self.schemes.keys())
        for scheme_id in ids:
            self.scheme_list.insert("end", self.schemes[scheme_id].get("name") or scheme_id)
        active_id = self.config_data.get("active_scheme_id") or DEFAULT_SCHEME_ID
        selected_index = ids.index(active_id) if active_id in ids else 0
        self.scheme_list.selection_set(selected_index)
        self.load_scheme(ids[selected_index])

    def on_scheme_select(self, _event=None):
        if not self.scheme_list.curselection():
            return
        self.save_current_scheme(show_message=False)
        scheme_id = list(self.schemes.keys())[self.scheme_list.curselection()[0]]
        self.load_scheme(scheme_id)

    def load_scheme(self, scheme_id):
        self.current_scheme_id = scheme_id
        scheme = self.schemes[scheme_id]
        self.name_var.set(scheme.get("name") or scheme_id)
        self.field_row_vars = []
        self.value_row_vars = []
        for widget in self.field_rows_frame.winfo_children():
            widget.destroy()
        for widget in self.value_rows_frame.winfo_children():
            widget.destroy()
        for item in self._scheme_field_items(scheme):
            self.add_field_row(item)
        for item in self._scheme_value_rules(scheme):
            self.add_value_row(item)
        self.sample_rate_var.set(str(scheme.get("invalid_sample_rate", 0.2)))
        self.sample_min_var.set(str(scheme.get("invalid_sample_min", 1)))
        self.overtime_var.set(str(scheme.get("overtime_threshold_minutes", 20)))
        self.company_text.delete("1.0", "end")
        self.company_text.insert("1.0", "\n".join(values.get("company_names") or sorted(DEFAULT_SCHEME["values"]["company_names"])))

    def save_current_scheme(self, show_message=True):
        if not self.current_scheme_id:
            return True
        field_items = self._collect_field_items()
        value_rules = self._collect_value_rules()
        fields = {}
        for item in field_items:
            if not item["enabled"]:
                continue
            if item["key"] == "special_targets":
                fields[item["key"]] = [part.strip() for part in item["column"].replace("，", ",").replace("、", ",").split(",") if part.strip()]
            else:
                fields[item["key"]] = item["column"]
        values = {item["key"]: item["value"] for item in value_rules if item["enabled"]}
        values["company_names"] = [line.strip() for line in self.company_text.get("1.0", "end").splitlines() if line.strip()]
        try:
            sample_rate = float(self.sample_rate_var.get() or 0.2)
            sample_min = int(self.sample_min_var.get() or 1)
            overtime = float(self.overtime_var.get() or 20)
        except ValueError:
            messagebox.showerror("配置错误", "抽样比例、最少条数和超时阈值必须是数字。")
            return False
        self.schemes[self.current_scheme_id] = {
            "id": self.current_scheme_id,
            "name": self.name_var.get().strip() or self.current_scheme_id,
            "fields": fields,
            "field_items": field_items,
            "values": values,
            "value_rules": value_rules,
            "invalid_sample_rate": sample_rate,
            "invalid_sample_min": sample_min,
            "overtime_threshold_minutes": overtime,
        }
        if show_message:
            messagebox.showinfo("已保存", "当前方案已暂存")
        return True

    def add_scheme(self):
        self.save_current_scheme(show_message=False)
        base = "scheme"
        idx = 1
        while f"{base}_{idx}" in self.schemes:
            idx += 1
        scheme_id = f"{base}_{idx}"
        scheme = copy_json(DEFAULT_SCHEME)
        scheme["id"] = scheme_id
        scheme["name"] = f"新方案{idx}"
        self.schemes[scheme_id] = scheme
        self.refresh_scheme_list()
        keys = list(self.schemes.keys())
        self.scheme_list.selection_clear(0, "end")
        self.scheme_list.selection_set(keys.index(scheme_id))
        self.load_scheme(scheme_id)

    def copy_scheme(self):
        if not self.current_scheme_id:
            return
        self.save_current_scheme(show_message=False)
        base_id = self.current_scheme_id
        idx = 1
        while f"{base_id}_copy_{idx}" in self.schemes:
            idx += 1
        scheme_id = f"{base_id}_copy_{idx}"
        scheme = copy_json(self.schemes[base_id])
        scheme["id"] = scheme_id
        scheme["name"] = scheme.get("name", base_id) + " 副本"
        self.schemes[scheme_id] = scheme
        self.refresh_scheme_list()

    def delete_scheme(self):
        if self.current_scheme_id == DEFAULT_SCHEME_ID:
            messagebox.showwarning("提示", "默认宏结构方案不能删除")
            return
        if not messagebox.askyesno("确认删除", "确认删除当前方案？"):
            return
        del self.schemes[self.current_scheme_id]
        if self.config_data.get("active_scheme_id") == self.current_scheme_id:
            self.config_data["active_scheme_id"] = DEFAULT_SCHEME_ID
        self.refresh_scheme_list()

    def reset_current_scheme(self):
        if not self.current_scheme_id:
            return
        default = copy_json(DEFAULT_SCHEME)
        default["id"] = self.current_scheme_id
        if self.current_scheme_id != DEFAULT_SCHEME_ID:
            default["name"] = self.name_var.get().strip() or self.current_scheme_id
        self.schemes[self.current_scheme_id] = default
        self.load_scheme(self.current_scheme_id)
        messagebox.showinfo("已恢复", "当前方案已恢复为默认宏结构")

    def set_active_scheme(self):
        self.save_current_scheme(show_message=False)
        self.config_data["active_scheme_id"] = self.current_scheme_id
        messagebox.showinfo("已设置", "当前方案已设为默认运行方案")

    def save_all_and_close(self):
        if not self.save_current_scheme(show_message=False):
            return
        self.config_data["schemes"] = self.schemes
        if not self.config_data.get("active_scheme_id"):
            self.config_data["active_scheme_id"] = self.current_scheme_id or DEFAULT_SCHEME_ID
        save_config(self.config_data)
        self.parent.refresh_schemes()
        self.parent.status.set("质检方案配置已保存")
        self.destroy()

    def load_headers(self):
        path = self.parent.input_path.get().strip()
        if not path or not Path(path).exists():
            path = filedialog.askopenfilename(
                title="选择用于读取表头的 Excel",
                filetypes=[("Excel 文件", "*.xlsx *.xlsm"), ("所有文件", "*.*")],
            )
        if not path:
            return
        try:
            wb = load_config_headers(path)
            self.header_choices = wb
        except Exception as exc:
            messagebox.showerror("读取失败", str(exc))
            return
        self._refresh_column_choices()
        messagebox.showinfo("完成", "已读取表头，可在字段映射中选择。")

    def _make_scroll_area(self, parent, row):
        canvas = tk.Canvas(parent, highlightthickness=0)
        scrollbar = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        frame = ttk.Frame(canvas)
        frame.bind("<Configure>", lambda _event: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas_window = canvas.create_window((0, 0), window=frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.grid(row=row, column=0, sticky="nsew")
        scrollbar.grid(row=row, column=1, sticky="ns")
        canvas.bind("<Configure>", lambda event: canvas.itemconfigure(canvas_window, width=event.width))
        return frame

    def _field_display(self, key):
        labels = dict(self.FIELD_OPTIONS)
        return f"{key} - {labels.get(key, key)}"

    def _value_display(self, key):
        labels = {item[0]: item[1] for item in self.VALUE_OPTIONS}
        return f"{key} - {labels.get(key, key)}"

    @staticmethod
    def _parse_key(display):
        return str(display or "").split(" - ", 1)[0].strip()

    def _column_choices(self):
        return self.header_choices or []

    def _scheme_field_items(self, scheme):
        items = copy_json(scheme.get("field_items") or [])
        if not items:
            fields = scheme.get("fields") or DEFAULT_SCHEME["fields"]
            labels = {item["key"]: item["label"] for item in DEFAULT_SCHEME["field_items"]}
            for key, column in fields.items():
                if isinstance(column, list):
                    column = ",".join(str(item) for item in column)
                items.append({"key": key, "label": labels.get(key, key), "column": str(column), "enabled": True})
        return items or copy_json(DEFAULT_SCHEME["field_items"])

    def _scheme_value_rules(self, scheme):
        items = copy_json(scheme.get("value_rules") or [])
        if not items:
            values = scheme.get("values") or DEFAULT_SCHEME["values"]
            for key, label, field_key, default_value in self.VALUE_OPTIONS:
                items.append({
                    "key": key,
                    "label": label,
                    "field_key": field_key,
                    "column": (scheme.get("fields") or DEFAULT_SCHEME["fields"]).get(field_key, ""),
                    "value": values.get(key, default_value),
                    "enabled": True,
                })
        return items or copy_json(DEFAULT_SCHEME["value_rules"])

    def add_field_row(self, item=None):
        item = item or {"key": "custom_field", "label": "自定义字段", "column": "", "enabled": True}
        row_index = len(self.field_row_vars)
        row = ttk.Frame(self.field_rows_frame, padding=(0, 3))
        row.grid(row=row_index, column=0, sticky="ew")
        row.columnconfigure(2, weight=1)
        row.columnconfigure(3, weight=2)
        enabled_var = tk.BooleanVar(value=item.get("enabled", True) is not False)
        key_var = tk.StringVar(value=self._field_display(item.get("key") or "custom_field"))
        label_var = tk.StringVar(value=item.get("label") or item.get("key") or "")
        column = item.get("column") or ""
        if isinstance(column, list):
            column = ",".join(str(part) for part in column)
        column_var = tk.StringVar(value=str(column))
        ttk.Checkbutton(row, variable=enabled_var).grid(row=0, column=0, padx=(0, 6))
        ttk.Combobox(row, textvariable=key_var, values=[self._field_display(key) for key, _ in self.FIELD_OPTIONS], width=22).grid(row=0, column=1, sticky="ew", padx=(0, 6))
        ttk.Entry(row, textvariable=label_var).grid(row=0, column=2, sticky="ew", padx=(0, 6))
        combo = ttk.Combobox(row, textvariable=column_var, values=self._column_choices())
        combo.grid(row=0, column=3, sticky="ew", padx=(0, 6))
        ttk.Button(row, text="删除", command=lambda: self._delete_dynamic_row(self.field_row_vars, row)).grid(row=0, column=4)
        self.field_row_vars.append({"frame": row, "enabled": enabled_var, "key": key_var, "label": label_var, "column": column_var, "combo": combo})

    def add_value_row(self, item=None):
        default_key, default_label, default_field_key, default_value = self.VALUE_OPTIONS[0]
        item = item or {"key": default_key, "label": default_label, "field_key": default_field_key, "column": "", "value": default_value, "enabled": True}
        row_index = len(self.value_row_vars)
        row = ttk.Frame(self.value_rows_frame, padding=(0, 3))
        row.grid(row=row_index, column=0, sticky="ew")
        row.columnconfigure(2, weight=1)
        row.columnconfigure(3, weight=2)
        row.columnconfigure(4, weight=1)
        enabled_var = tk.BooleanVar(value=item.get("enabled", True) is not False)
        key_var = tk.StringVar(value=self._value_display(item.get("key") or default_key))
        label_var = tk.StringVar(value=item.get("label") or item.get("key") or "")
        column_var = tk.StringVar(value=str(item.get("column") or ""))
        value_var = tk.StringVar(value=str(item.get("value") or ""))
        ttk.Checkbutton(row, variable=enabled_var).grid(row=0, column=0, padx=(0, 6))
        ttk.Combobox(row, textvariable=key_var, values=[self._value_display(key) for key, _label, _field_key, _value in self.VALUE_OPTIONS], width=24).grid(row=0, column=1, sticky="ew", padx=(0, 6))
        ttk.Entry(row, textvariable=label_var).grid(row=0, column=2, sticky="ew", padx=(0, 6))
        combo = ttk.Combobox(row, textvariable=column_var, values=self._column_choices())
        combo.grid(row=0, column=3, sticky="ew", padx=(0, 6))
        ttk.Entry(row, textvariable=value_var).grid(row=0, column=4, sticky="ew", padx=(0, 6))
        ttk.Button(row, text="删除", command=lambda: self._delete_dynamic_row(self.value_row_vars, row)).grid(row=0, column=5)
        self.value_row_vars.append({"frame": row, "enabled": enabled_var, "key": key_var, "label": label_var, "column": column_var, "value": value_var, "combo": combo})

    def _delete_dynamic_row(self, rows, frame):
        for idx, item in enumerate(list(rows)):
            if item["frame"] == frame:
                item["frame"].destroy()
                del rows[idx]
                break

    def _refresh_column_choices(self):
        for item in self.field_row_vars + self.value_row_vars:
            item["combo"]["values"] = self._column_choices()

    def _collect_field_items(self):
        items = []
        for row in self.field_row_vars:
            key = self._parse_key(row["key"].get())
            label = row["label"].get().strip() or key
            column = row["column"].get().strip()
            if not key and not column:
                continue
            items.append({"key": key, "label": label, "column": column, "enabled": bool(row["enabled"].get())})
        return items

    def _collect_value_rules(self):
        field_map = {key: field_key for key, _label, field_key, _default in self.VALUE_OPTIONS}
        items = []
        for row in self.value_row_vars:
            key = self._parse_key(row["key"].get())
            label = row["label"].get().strip() or key
            column = row["column"].get().strip()
            value = row["value"].get().strip()
            if not key and not value:
                continue
            items.append({
                "key": key,
                "label": label,
                "field_key": field_map.get(key, ""),
                "column": column,
                "value": value,
                "enabled": bool(row["enabled"].get()),
            })
        return items


def load_config_headers(path):
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=False)
    try:
        ws = wb.worksheets[0]
        row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        choices = []
        for idx, value in enumerate(row, start=1):
            text = str(value).strip() if value is not None else ""
            letter = get_column_name(idx)
            choices.append(f"{letter} {text}" if text else letter)
        return choices
    finally:
        wb.close()


def get_column_name(index):
    name = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        name = chr(65 + remainder) + name
    return name


class SpecialConfigWindow(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent
        self.title("专项质检配置")
        self.geometry("780x560")
        self.minsize(720, 500)
        self.transient(parent)

        self.config_data = load_config()
        self.plans = copy.deepcopy(self.config_data.get("monthly_special_plans") or get_default_monthly_plans())
        self.current_month = tk.StringVar(value="1")
        self.name_var = tk.StringVar()

        self._build_ui()
        self.month_list.selection_set(0)
        self.load_month("1")

    def _build_ui(self):
        self.columnconfigure(1, weight=1)
        self.rowconfigure(0, weight=1)

        left = ttk.Frame(self, padding=(14, 14))
        left.grid(row=0, column=0, sticky="ns")
        ttk.Label(left, text="月份").pack(anchor="w")
        self.month_list = tk.Listbox(left, height=12, exportselection=False, width=12)
        self.month_list.pack(fill="y", expand=True, pady=(8, 0))
        for month in range(1, 13):
            self.month_list.insert("end", f"{month}月")
        self.month_list.bind("<<ListboxSelect>>", self.on_month_select)

        right = ttk.Frame(self, padding=(0, 14, 14, 14))
        right.grid(row=0, column=1, sticky="nsew")
        right.columnconfigure(0, weight=1)
        right.rowconfigure(5, weight=1)

        ttk.Label(right, text="策略名称").grid(row=0, column=0, sticky="w")
        ttk.Entry(right, textvariable=self.name_var).grid(row=1, column=0, sticky="ew", pady=(6, 12))

        ttk.Label(right, text="质检策略").grid(row=2, column=0, sticky="w")
        self.strategy_text = scrolledtext.ScrolledText(right, height=5, wrap="word")
        self.strategy_text.grid(row=3, column=0, sticky="ew", pady=(6, 12))

        ttk.Label(right, text="质检关键词（用逗号、顿号或换行分隔）").grid(row=4, column=0, sticky="w")
        self.keyword_text = scrolledtext.ScrolledText(right, height=12, wrap="word")
        self.keyword_text.grid(row=5, column=0, sticky="nsew", pady=(6, 12))

        actions = ttk.Frame(right)
        actions.grid(row=6, column=0, sticky="ew")
        ttk.Button(actions, text="保存当前月份", command=self.save_current_month).pack(side="left")
        ttk.Button(actions, text="恢复当前月份默认值", command=self.reset_current_month).pack(side="left", padx=8)
        ttk.Button(actions, text="保存全部配置", command=self.save_all).pack(side="right")

    def on_month_select(self, _event=None):
        if not self.month_list.curselection():
            return
        self.save_current_month(show_message=False)
        month = str(self.month_list.curselection()[0] + 1)
        self.load_month(month)

    def load_month(self, month):
        self.current_month.set(month)
        plan = self.plans.get(month) or get_default_monthly_plans()[month]
        self.name_var.set(plan.get("name", ""))
        self.strategy_text.delete("1.0", "end")
        self.strategy_text.insert("1.0", plan.get("strategy", ""))
        self.keyword_text.delete("1.0", "end")
        self.keyword_text.insert("1.0", "，".join(plan.get("keywords") or []))

    def save_current_month(self, show_message=True):
        month = self.current_month.get()
        keywords = self._parse_keywords(self.keyword_text.get("1.0", "end"))
        self.plans[month] = {
            "name": self.name_var.get().strip() or f"{month}月专项",
            "strategy": self.strategy_text.get("1.0", "end").strip(),
            "keywords": keywords,
        }
        if show_message:
            messagebox.showinfo("已保存", f"{month}月专项配置已暂存")

    def reset_current_month(self):
        month = self.current_month.get()
        self.plans[month] = get_default_monthly_plans()[month]
        self.load_month(month)
        messagebox.showinfo("已恢复", f"{month}月专项配置已恢复默认值")

    def save_all(self):
        self.save_current_month(show_message=False)
        self.config_data["monthly_special_plans"] = self.plans
        save_config(self.config_data)
        self.parent.status.set("专项质检配置已保存")
        messagebox.showinfo("完成", "专项质检配置已保存，下一次处理立即生效。")

    @staticmethod
    def _parse_keywords(text):
        normalized = text.replace("，", ",").replace("、", ",").replace("\n", ",")
        return [item.strip() for item in normalized.split(",") if item.strip()]


if __name__ == "__main__":
    app = ReviewTool()
    app.mainloop()
