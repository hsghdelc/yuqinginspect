import random
import json
import sys
import shutil
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from monthly_special import get_default_monthly_plans, get_monthly_plan


COMPANIES = {
    "国网甘肃电力", "国网河北电力", "国网黑龙江电力", "国网湖北电力", "国网江苏电力",
    "国网蒙东电力", "国网山东电力", "国网陕西电力", "国网上海电力", "国网四川电力",
    "国网西藏电力", "国网浙江电力", "国网重庆电力", "国网冀北电力",
}


DEFAULT_SCHEME_ID = "default"

DEFAULT_SCHEME = {
    "id": DEFAULT_SCHEME_ID,
    "name": "默认结构方案",
    "fields": {
        "keep": "AG",
        "company": "H",
        "send_time": "A",
        "process_time": "AB",
        "event_nature": "O",
        "valid_scope": "R",
        "marketing": "S",
        "reminder": "T",
        "duplicate": "U",
        "event_category": "Y",
        "id": "B",
        "special_targets": ["C", "K"],
    },
    "field_items": [
        {"key": "keep", "label": "基础保留列", "column": "AG", "enabled": True},
        {"key": "company", "label": "公司筛选列", "column": "H", "enabled": True},
        {"key": "send_time", "label": "工单派发时间列", "column": "A", "enabled": True},
        {"key": "process_time", "label": "客服部处理时间列", "column": "AB", "enabled": True},
        {"key": "event_nature", "label": "事件性质列", "column": "O", "enabled": True},
        {"key": "valid_scope", "label": "是否符合舆情范围列", "column": "R", "enabled": True},
        {"key": "marketing", "label": "是否营销类舆情事件列", "column": "S", "enabled": True},
        {"key": "reminder", "label": "舆情提醒标识列", "column": "T", "enabled": True},
        {"key": "duplicate", "label": "是否重复事件列", "column": "U", "enabled": True},
        {"key": "event_category", "label": "事件分类列", "column": "Y", "enabled": True},
        {"key": "id", "label": "编号列", "column": "B", "enabled": True},
        {"key": "special_targets", "label": "专项关键词匹配列", "column": "C,K", "enabled": True},
    ],
    "values": {
        "company_names": sorted(COMPANIES),
        "valid_yes": "是",
        "valid_no": "否",
        "marketing_yes": "是",
        "marketing_no": "否",
        "duplicate_yes": "是",
        "duplicate_no": "否",
        "negative_event": "负面事件",
        "positive_event": "正面事件",
        "reminder": "舆情提醒",
        "livelihood_event": "民生类舆情",
    },
    "value_rules": [
        {"key": "valid_yes", "label": "符合舆情范围：是", "field_key": "valid_scope", "column": "R", "value": "是", "enabled": True},
        {"key": "valid_no", "label": "符合舆情范围：否", "field_key": "valid_scope", "column": "R", "value": "否", "enabled": True},
        {"key": "marketing_yes", "label": "营销类：是", "field_key": "marketing", "column": "S", "value": "是", "enabled": True},
        {"key": "marketing_no", "label": "营销类：否", "field_key": "marketing", "column": "S", "value": "否", "enabled": True},
        {"key": "duplicate_yes", "label": "重复事件：是", "field_key": "duplicate", "column": "U", "value": "是", "enabled": True},
        {"key": "duplicate_no", "label": "重复事件：否", "field_key": "duplicate", "column": "U", "value": "否", "enabled": True},
        {"key": "negative_event", "label": "负面事件值", "field_key": "event_nature", "column": "O", "value": "负面事件", "enabled": True},
        {"key": "positive_event", "label": "正面事件值", "field_key": "event_nature", "column": "O", "value": "正面事件", "enabled": True},
        {"key": "reminder", "label": "舆情提醒值", "field_key": "reminder", "column": "T", "value": "舆情提醒", "enabled": True},
        {"key": "livelihood_event", "label": "民生事件值", "field_key": "event_category", "column": "Y", "value": "民生类舆情", "enabled": True},
    ],
    "review_plans": [
        {
            "name": "月度专项质检",
            "role": "monthly_special",
            "enabled": True,
            "output_sheet": True,
            "match_type": "按月份专项关键词",
            "keyword_columns": "C,K",
            "keywords": "",
            "match_mode": "包含",
            "case_sensitive": False,
            "apply_conditions": False,
            "conditions": [],
            "sampling": {"enabled": False, "mode": "按比例", "value": 0.2, "min_count": 1},
            "overtime": {"enabled": False, "mode": "按起止时间计算", "send_column": "A", "process_column": "AB", "duration_column": "", "threshold_minutes": 20, "id_column": "B"},
        },
        {
            "name": "无效复核",
            "role": "invalid_review",
            "enabled": True,
            "output_sheet": True,
            "match_type": "条件筛选",
            "keyword_columns": "",
            "keywords": "",
            "match_mode": "包含",
            "case_sensitive": False,
            "apply_conditions": False,
            "conditions": [{"column": "R", "operator": "等于", "value": "否"}],
            "sampling": {"enabled": True, "mode": "按比例", "value": 0.2, "min_count": 1},
            "overtime": {"enabled": False, "mode": "按起止时间计算", "send_column": "A", "process_column": "AB", "duration_column": "", "threshold_minutes": 20, "id_column": "B"},
        },
        {
            "name": "舆情提醒复核",
            "role": "reminder_review",
            "enabled": True,
            "output_sheet": True,
            "match_type": "条件筛选",
            "keyword_columns": "",
            "keywords": "",
            "match_mode": "包含",
            "case_sensitive": False,
            "apply_conditions": False,
            "conditions": [{"column": "T", "operator": "等于", "value": "舆情提醒"}],
            "sampling": {"enabled": False, "mode": "按比例", "value": 0.2, "min_count": 1},
            "overtime": {"enabled": False, "mode": "按起止时间计算", "send_column": "A", "process_column": "AB", "duration_column": "", "threshold_minutes": 20, "id_column": "B"},
        },
        {
            "name": "超时检查",
            "role": "overtime_check",
            "enabled": True,
            "output_sheet": False,
            "match_type": "条件筛选",
            "keyword_columns": "",
            "keywords": "",
            "match_mode": "包含",
            "case_sensitive": False,
            "apply_conditions": False,
            "conditions": [
                {"column": "R", "operator": "等于", "value": "是"},
                {"column": "S", "operator": "等于", "value": "是"},
                {"column": "U", "operator": "等于", "value": "否"},
            ],
            "sampling": {"enabled": False, "mode": "按比例", "value": 0.2, "min_count": 1},
            "overtime": {"enabled": True, "mode": "按起止时间计算", "send_column": "A", "process_column": "AB", "duration_column": "", "threshold_minutes": 20, "id_column": "B"},
        },
    ],
    "invalid_sample_rate": 0.2,
    "invalid_sample_min": 1,
    "overtime_threshold_minutes": 20,
}


def _app_dir():
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def load_config(config_path=None):
    path = Path(config_path) if config_path else _app_dir() / "config.json"
    default = {
        "companies": sorted(COMPANIES),
        "special_target_columns": [3, 11],
        "overtime_threshold_minutes": 20,
        "preserve_cell_styles": False,
        "monthly_special_plans": get_default_monthly_plans(),
        "active_scheme_id": DEFAULT_SCHEME_ID,
        "schemes": {DEFAULT_SCHEME_ID: DEFAULT_SCHEME},
        "config_version": datetime.now().strftime("%Y-%m-%d-001"),
        "config_remark": "",
    }
    if not path.exists():
        return default
    with path.open("r", encoding="utf-8") as f:
        loaded = json.load(f)
    default.update({k: v for k, v in loaded.items() if v is not None})
    default["schemes"] = _normalize_schemes(default)
    return default


def save_config(config, config_path=None):
    path = Path(config_path) if config_path else _app_dir() / "config.json"
    if config_path is None and path.exists():
        backup_dir = _app_dir() / "config_backups"
        backup_dir.mkdir(exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        shutil.copy2(path, backup_dir / f"配置备份_{stamp}.json")
    with path.open("w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=2)


def _normalize_schemes(config):
    schemes = copy_json(config.get("schemes")) if config.get("schemes") else {}
    if not isinstance(schemes, dict):
        schemes = {}
    if DEFAULT_SCHEME_ID not in schemes:
        schemes[DEFAULT_SCHEME_ID] = copy_json(DEFAULT_SCHEME)
    for scheme_id, scheme in list(schemes.items()):
        merged = copy_json(DEFAULT_SCHEME)
        merged.update(scheme or {})
        merged["id"] = scheme_id
        if scheme_id == DEFAULT_SCHEME_ID and merged.get("name") == "默认宏结构方案":
            merged["name"] = DEFAULT_SCHEME["name"]
        merged["fields"] = {**DEFAULT_SCHEME["fields"], **((scheme or {}).get("fields") or {})}
        merged["values"] = {**DEFAULT_SCHEME["values"], **((scheme or {}).get("values") or {})}
        merged["field_items"] = _normalize_field_items(merged)
        merged["value_rules"] = _normalize_value_rules(merged)
        merged["review_plans"] = _normalize_review_plans(merged)
        schemes[scheme_id] = merged
    return schemes


def _normalize_field_items(scheme):
    field_items = scheme.get("field_items")
    if not isinstance(field_items, list) or not field_items:
        field_items = []
        fields = scheme.get("fields") or {}
        labels = {item["key"]: item["label"] for item in DEFAULT_SCHEME["field_items"]}
        for key, column in fields.items():
            if isinstance(column, list):
                column = ",".join(str(item) for item in column)
            field_items.append({
                "key": key,
                "label": labels.get(key, key),
                "column": str(column),
                "enabled": True,
            })
    known = {item.get("key"): item for item in field_items if isinstance(item, dict)}
    for default_item in DEFAULT_SCHEME["field_items"]:
        if default_item["key"] not in known:
            field_items.append(copy_json(default_item))
    return [
        {
            "key": str(item.get("key") or "").strip(),
            "label": str(item.get("label") or item.get("key") or "").strip(),
            "column": str(item.get("column") or "").strip(),
            "enabled": item.get("enabled", True) is not False,
        }
        for item in field_items
        if isinstance(item, dict) and str(item.get("key") or "").strip()
    ]


def _normalize_value_rules(scheme):
    value_rules = scheme.get("value_rules")
    if not isinstance(value_rules, list) or not value_rules:
        value_rules = []
        values = scheme.get("values") or {}
        fields = _fields_from_items(scheme.get("field_items") or DEFAULT_SCHEME["field_items"])
        defaults = {item["key"]: item for item in DEFAULT_SCHEME["value_rules"]}
        for key, default_rule in defaults.items():
            rule = copy_json(default_rule)
            rule["value"] = values.get(key, default_rule["value"])
            rule["column"] = fields.get(default_rule["field_key"], default_rule["column"])
            value_rules.append(rule)
    known = {item.get("key"): item for item in value_rules if isinstance(item, dict)}
    for default_rule in DEFAULT_SCHEME["value_rules"]:
        if default_rule["key"] not in known:
            value_rules.append(copy_json(default_rule))
    return [
        {
            "key": str(item.get("key") or "").strip(),
            "label": str(item.get("label") or item.get("key") or "").strip(),
            "field_key": str(item.get("field_key") or "").strip(),
            "column": str(item.get("column") or "").strip(),
            "value": str(item.get("value") or "").strip(),
            "enabled": item.get("enabled", True) is not False,
        }
        for item in value_rules
        if isinstance(item, dict) and str(item.get("key") or "").strip()
    ]


def _fields_from_items(field_items):
    fields = {}
    for item in field_items or []:
        if not isinstance(item, dict) or item.get("enabled", True) is False:
            continue
        key = str(item.get("key") or "").strip()
        column = str(item.get("column") or "").strip()
        if not key or not column:
            continue
        if key == "special_targets":
            fields[key] = [part.strip() for part in column.replace("，", ",").replace("、", ",").split(",") if part.strip()]
        else:
            fields[key] = column
    return fields


def _values_from_rules(value_rules):
    values = {}
    field_overrides = {}
    for rule in value_rules or []:
        if not isinstance(rule, dict) or rule.get("enabled", True) is False:
            continue
        key = str(rule.get("key") or "").strip()
        if not key:
            continue
        values[key] = str(rule.get("value") or "").strip()
        field_key = str(rule.get("field_key") or "").strip()
        column = str(rule.get("column") or "").strip()
        if field_key and column:
            field_overrides[field_key] = column
    return values, field_overrides


def _normalize_review_plans(scheme):
    plans = scheme.get("review_plans")
    if isinstance(plans, list) and plans:
        return [_normalize_review_plan(plan) for plan in plans if isinstance(plan, dict)]
    fields = {**DEFAULT_SCHEME["fields"], **(scheme.get("fields") or {}), **_fields_from_items(scheme.get("field_items"))}
    values = {**DEFAULT_SCHEME["values"], **(scheme.get("values") or {})}
    defaults = copy_json(DEFAULT_SCHEME["review_plans"])
    for plan in defaults:
        if plan.get("role") == "monthly_special":
            columns = fields.get("special_targets") or ["C", "K"]
            plan["keyword_columns"] = ",".join(columns) if isinstance(columns, list) else str(columns)
        elif plan.get("role") == "invalid_review":
            plan["conditions"] = [{"column": fields.get("valid_scope", "R"), "operator": "等于", "value": values.get("valid_no", "否")}]
            plan["sampling"] = {
                "enabled": True,
                "mode": "按比例",
                "value": float(scheme.get("invalid_sample_rate", 0.2) or 0.2),
                "min_count": int(scheme.get("invalid_sample_min", 1) or 1),
            }
        elif plan.get("role") == "reminder_review":
            plan["conditions"] = [{"column": fields.get("reminder", "T"), "operator": "等于", "value": values.get("reminder", "舆情提醒")}]
        elif plan.get("role") == "overtime_check":
            plan["conditions"] = [
                {"column": fields.get("valid_scope", "R"), "operator": "等于", "value": values.get("valid_yes", "是")},
                {"column": fields.get("marketing", "S"), "operator": "等于", "value": values.get("marketing_yes", "是")},
                {"column": fields.get("duplicate", "U"), "operator": "等于", "value": values.get("duplicate_no", "否")},
            ]
            plan["overtime"] = {
                "enabled": True,
                "mode": "按起止时间计算",
                "send_column": fields.get("send_time", "A"),
                "process_column": fields.get("process_time", "AB"),
                "duration_column": "",
                "threshold_minutes": float(scheme.get("overtime_threshold_minutes", 20) or 20),
                "id_column": fields.get("id", "B"),
            }
    return [_normalize_review_plan(plan) for plan in defaults]


def _normalize_review_plan(plan):
    sampling = plan.get("sampling") if isinstance(plan.get("sampling"), dict) else {}
    overtime = plan.get("overtime") if isinstance(plan.get("overtime"), dict) else {}
    return {
        "name": str(plan.get("name") or "质检计划").strip(),
        "role": str(plan.get("role") or "").strip(),
        "enabled": plan.get("enabled", True) is not False,
        "output_sheet": plan.get("output_sheet", True) is not False,
        "match_type": str(plan.get("match_type") or "条件筛选").strip(),
        "keyword_columns": plan.get("keyword_columns") or "",
        "keywords": str(plan.get("keywords") or ""),
        "match_mode": str(plan.get("match_mode") or "包含").strip(),
        "case_sensitive": bool(plan.get("case_sensitive", False)),
        "apply_conditions": plan.get("apply_conditions", False) is True,
        "conditions": [
            {
                "column": str(cond.get("column") or "").strip(),
                "operator": str(cond.get("operator") or "等于").strip(),
                "value": str(cond.get("value") or "").strip(),
            }
            for cond in (plan.get("conditions") or [])
            if isinstance(cond, dict)
        ],
        "sampling": {
            "enabled": sampling.get("enabled", False) is True,
            "mode": str(sampling.get("mode") or "按比例").strip(),
            "value": sampling.get("value", 0.2),
            "min_count": sampling.get("min_count", 1),
        },
        "overtime": {
            "enabled": overtime.get("enabled", False) is True,
            "mode": str(overtime.get("mode") or "按起止时间计算").strip(),
            "send_column": str(overtime.get("send_column") or "A").strip(),
            "process_column": str(overtime.get("process_column") or "AB").strip(),
            "duration_column": str(overtime.get("duration_column") or "").strip(),
            "threshold_minutes": overtime.get("threshold_minutes", 20),
            "id_column": str(overtime.get("id_column") or "B").strip(),
        },
    }


def copy_json(value):
    return json.loads(json.dumps(value, ensure_ascii=False))


def get_active_scheme(config, scheme_id=None):
    schemes = _normalize_schemes(config)
    selected_id = scheme_id or config.get("active_scheme_id") or DEFAULT_SCHEME_ID
    if selected_id not in schemes:
        selected_id = DEFAULT_SCHEME_ID
    return schemes[selected_id]


def _cell_text(value):
    return "" if value is None else str(value).strip()


def _safe_sheet_name(name):
    for ch in ':\\/?*[]':
        name = name.replace(ch, "")
    return name[:31]


def _find_col(headers, contains, default_index):
    for idx, value in enumerate(headers, start=1):
        if contains in _cell_text(value):
            return idx
    return default_index


def _letters_to_col(text):
    value = 0
    for ch in text.upper():
        if not ("A" <= ch <= "Z"):
            return None
        value = value * 26 + ord(ch) - ord("A") + 1
    return value or None


def _resolve_col(spec, headers, default_index):
    if spec is None or spec == "":
        return default_index
    text = str(spec).strip()
    if not text:
        return default_index
    first_token = text.split()[0] if text.split() else text
    token_letter_col = _letters_to_col(first_token)
    if token_letter_col:
        return token_letter_col
    if text.isdigit():
        return int(text)
    letter_col = _letters_to_col(text)
    if letter_col:
        return letter_col
    for idx, value in enumerate(headers, start=1):
        if _cell_text(value) == text:
            return idx
    for idx, value in enumerate(headers, start=1):
        if text in _cell_text(value):
            return idx
    return default_index


def _resolve_cols(specs, headers, default_indexes):
    if isinstance(specs, str):
        specs = [item.strip() for item in specs.replace("，", ",").replace("、", ",").split(",") if item.strip()]
    if not specs:
        return default_indexes
    columns = []
    for idx, spec in enumerate(specs):
        default_index = default_indexes[min(idx, len(default_indexes) - 1)]
        col = _resolve_col(spec, headers, default_index)
        if col > 0 and col not in columns:
            columns.append(col)
    return columns or default_indexes


def _parse_col_specs(specs):
    if isinstance(specs, (list, tuple)):
        parts = specs
    else:
        parts = str(specs or "").replace("，", ",").replace("、", ",").split(",")
    return [str(part).strip() for part in parts if str(part).strip()]


def _spec_exists(spec, headers):
    text = str(spec or "").strip()
    if not text:
        return False
    col = _resolve_col(text, headers, 0)
    return 1 <= col <= len(headers)


def _validate_column_specs(label, specs, headers, errors):
    parts = _parse_col_specs(specs)
    if not parts:
        errors.append(f"{label}未配置列。")
        return
    for part in parts:
        if not _spec_exists(part, headers):
            errors.append(f"{label}配置为【{part}】，当前源文件未识别到该列。")


def load_workbook_headers(path):
    wb = load_workbook(path, read_only=True, data_only=False)
    try:
        ws = wb.worksheets[0]
        if ws.max_row == 1 and ws.max_column == 1:
            ws.reset_dimensions()
        row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        headers = list(row)
        max_col = max(ws.max_column or 0, len(headers))
        max_row = ws.max_row
        return headers, max_row, max_col
    finally:
        wb.close()


def validate_config_for_file(input_path=None, scheme_id=None, config_path=None):
    config = load_config(config_path)
    scheme = get_active_scheme(config, scheme_id)
    errors = []
    warnings = []
    headers = []
    max_row = None
    max_col = 0
    if input_path:
        try:
            headers, max_row, max_col = load_workbook_headers(input_path)
            if not headers:
                errors.append("源文件未读取到表头。")
        except Exception as exc:
            errors.append(f"源文件表头读取失败：{exc}")

    fields = {**(scheme.get("fields") or DEFAULT_SCHEME["fields"]), **_fields_from_items(scheme.get("field_items"))}
    values = scheme.get("values") or {}
    if not values.get("company_names"):
        warnings.append("公司名单为空，公司筛选可能无法保留任何数据。")

    if headers:
        for label, spec in [("基础保留列", fields.get("keep")), ("公司筛选列", fields.get("company")), ("编号列", fields.get("id"))]:
            _validate_column_specs(label, spec, headers, errors)

    active_plans = [plan for plan in _normalize_review_plans(scheme) if plan.get("enabled", True)]
    if not active_plans:
        errors.append("当前方案未启用任何质检计划。")
    for idx, plan in enumerate(active_plans, start=1):
        name = plan.get("name") or f"质检计划{idx}"
        match_type = plan.get("match_type", "条件筛选")
        if match_type in {"按月份专项关键词", "关键词筛选"}:
            if headers:
                _validate_column_specs(f"质检计划【{name}】关键词匹配列", plan.get("keyword_columns"), headers, errors)
            if match_type == "关键词筛选" and not _split_items(plan.get("keywords")):
                errors.append(f"质检计划【{name}】选择了关键词筛选，但关键词为空。")
        needs_conditions = match_type == "条件筛选" or plan.get("apply_conditions")
        if needs_conditions and not plan.get("conditions"):
            errors.append(f"质检计划【{name}】需要条件筛选，但条件为空。")
        if headers:
            for condition in plan.get("conditions") or []:
                col = condition.get("column")
                if not _spec_exists(col, headers):
                    errors.append(f"质检计划【{name}】条件列【{str(col or '').strip() or '未填写'}】当前源文件未识别到。")
        sampling = plan.get("sampling") or {}
        if sampling.get("enabled"):
            try:
                value = float(sampling.get("value", 0))
                min_count = int(sampling.get("min_count", 1))
                if value <= 0:
                    errors.append(f"质检计划【{name}】抽样值必须大于 0。")
                if min_count < 0:
                    errors.append(f"质检计划【{name}】最少条数不能小于 0。")
                if sampling.get("mode") == "按比例" and value > 100:
                    warnings.append(f"质检计划【{name}】抽样比例超过 100%，将等同于保留全部命中数据。")
            except (TypeError, ValueError):
                errors.append(f"质检计划【{name}】抽样值或最少条数不是有效数字。")
        overtime = plan.get("overtime") or {}
        if overtime.get("enabled"):
            try:
                threshold = float(overtime.get("threshold_minutes", 20))
                if threshold < 0:
                    errors.append(f"质检计划【{name}】超时阈值不能小于 0。")
            except (TypeError, ValueError):
                errors.append(f"质检计划【{name}】超时阈值不是有效数字。")
            if headers:
                _validate_column_specs(f"质检计划【{name}】编号列", overtime.get("id_column"), headers, errors)
                if overtime.get("mode") == "使用已有处理时长列":
                    _validate_column_specs(f"质检计划【{name}】已有处理时长列", overtime.get("duration_column"), headers, errors)
                else:
                    _validate_column_specs(f"质检计划【{name}】派发时间列", overtime.get("send_column"), headers, errors)
                    _validate_column_specs(f"质检计划【{name}】处理时间列", overtime.get("process_column"), headers, errors)

    return {
        "ok": not errors,
        "errors": errors,
        "warnings": warnings,
        "scheme_name": scheme.get("name", DEFAULT_SCHEME["name"]),
        "header_count": len(headers),
        "row_count": (max_row - 1) if isinstance(max_row, int) and max_row > 0 else None,
        "column_count": max_col,
    }


def _row_value(row, col):
    idx = col - 1
    if idx < 0 or idx >= len(row):
        return None
    return row[idx]


def _row_text(row, col):
    return _cell_text(_row_value(row, col))


def _set_row_value(row, col, value):
    values = list(row)
    while len(values) < col:
        values.append(None)
    values[col - 1] = value
    return tuple(values)


def _parse_datetime(value):
    if isinstance(value, datetime):
        return value
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    try:
        return datetime.fromisoformat(text.replace("/", "-"))
    except ValueError:
        return None


def _conditions_to_text(conditions):
    parts = []
    for condition in conditions or []:
        column = str(condition.get("column") or "").strip()
        value = str(condition.get("value") or "").strip()
        operator = str(condition.get("operator") or "等于").strip()
        if not column and not value:
            continue
        if operator == "等于":
            parts.append(f"{column}={value}")
        elif operator == "不等于":
            parts.append(f"{column}!={value}")
        else:
            parts.append(f"{column}{operator}{value}")
    return ";".join(parts)


def _matched_keywords(row, plan, headers, month_plan=None):
    columns = _resolve_cols(plan.get("keyword_columns"), headers, [3, 11])
    source_keywords = month_plan.get("keywords") if month_plan else plan.get("keywords")
    original_keywords = _split_items(source_keywords)
    if not original_keywords:
        return []
    text = " ".join(_row_text(row, col) for col in columns)
    compare_text = text if plan.get("case_sensitive", False) else text.lower()
    matches = []
    for keyword in original_keywords:
        compare_keyword = keyword if plan.get("case_sensitive", False) else keyword.lower()
        if plan.get("match_mode") == "等于":
            hit = compare_text == compare_keyword
        else:
            hit = compare_keyword in compare_text
        if hit:
            matches.append(keyword)
    return matches


def _hit_reason(plan, row, headers, duration_col, month_plan=None):
    match_type = plan.get("match_type", "条件筛选")
    has_keyword_match = match_type in {"按月份专项关键词", "关键词筛选"}
    keywords = _matched_keywords(row, plan, headers, month_plan) if has_keyword_match else []
    condition_text = _conditions_to_text(plan.get("conditions") or []) if (match_type == "条件筛选" or plan.get("apply_conditions")) else ""
    mode = "关键词筛选 + 条件筛选" if has_keyword_match and plan.get("apply_conditions") else match_type
    overtime = plan.get("overtime") or {}
    duration = _row_text(row, duration_col)
    if overtime.get("enabled"):
        try:
            overtime_text = "超时" if float(duration) > float(overtime.get("threshold_minutes", 20) or 20) else "未超时"
        except (TypeError, ValueError):
            overtime_text = "无法判断"
    else:
        overtime_text = "未启用"
    return {
        "plan": plan.get("name", "质检计划"),
        "mode": mode,
        "keywords": "，".join(keywords),
        "conditions": condition_text,
        "overtime": overtime_text,
    }


def _append_review_sheet(wb, title, header, rows, inspector, plan=None, source_headers=None, duration_col=None, month_plan=None):
    ws = wb.create_sheet(title)
    base_col = len(header)
    header = _set_row_value(header, base_col + 1, "质检人员")
    header = _set_row_value(header, base_col + 2, "质检结果")
    header = _set_row_value(header, base_col + 3, "补录单号")
    header = _set_row_value(header, base_col + 4, "命中质检计划")
    header = _set_row_value(header, base_col + 5, "命中方式")
    header = _set_row_value(header, base_col + 6, "命中关键词")
    header = _set_row_value(header, base_col + 7, "命中条件")
    header = _set_row_value(header, base_col + 8, "超时判断")
    ws.append(header)
    for row in rows:
        reason = _hit_reason(plan or {}, row, source_headers or [], duration_col or len(header), month_plan)
        row = _set_row_value(row, base_col + 1, inspector)
        row = _set_row_value(row, base_col + 2, "通过")
        row = _set_row_value(row, base_col + 3, "")
        row = _set_row_value(row, base_col + 4, reason["plan"])
        row = _set_row_value(row, base_col + 5, reason["mode"])
        row = _set_row_value(row, base_col + 6, reason["keywords"])
        row = _set_row_value(row, base_col + 7, reason["conditions"])
        row = _set_row_value(row, base_col + 8, reason["overtime"])
        ws.append(row)
    for col in range(1, len(header) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15


def _count(rows, col, value):
    return sum(1 for row in rows if _row_text(row, col) == value)


def _countifs(rows, pairs):
    return sum(1 for row in rows if all(_row_text(row, col) == value for col, value in pairs))


def _split_items(text):
    if isinstance(text, list):
        return [str(item).strip() for item in text if str(item).strip()]
    return [item.strip() for item in str(text or "").replace("，", ",").replace("、", ",").replace("\n", ",").split(",") if item.strip()]


def _match_text(value, expected, operator):
    actual = _cell_text(value)
    expected = str(expected or "").strip()
    if operator == "包含":
        return expected in actual
    if operator == "不包含":
        return expected not in actual
    if operator == "不等于":
        return actual != expected
    if operator == "为空":
        return actual == ""
    if operator == "非空":
        return actual != ""
    return actual == expected


def _match_conditions(row, conditions, headers):
    if not conditions:
        return True
    for condition in conditions:
        col = _resolve_col(condition.get("column"), headers, 1)
        if not _match_text(_row_value(row, col), condition.get("value"), condition.get("operator", "等于")):
            return False
    return True


def _match_keywords(row, plan, headers, month_plan=None):
    return bool(_matched_keywords(row, plan, headers, month_plan))


def _run_review_plan(rows, plan, headers, month_plan=None):
    match_type = plan.get("match_type", "条件筛选")
    if match_type == "按月份专项关键词":
        matched = [row for row in rows if _match_keywords(row, plan, headers, month_plan)]
    elif match_type == "关键词筛选":
        matched = [row for row in rows if _match_keywords(row, plan, headers)]
    else:
        matched = [row for row in rows if _match_conditions(row, plan.get("conditions") or [], headers)]
    if match_type in {"按月份专项关键词", "关键词筛选"} and plan.get("apply_conditions"):
        matched = [row for row in matched if _match_conditions(row, plan.get("conditions") or [], headers)]
    sampling = plan.get("sampling") or {}
    if sampling.get("enabled") and matched:
        mode = sampling.get("mode", "按比例")
        value = float(sampling.get("value", 0.2) or 0.2)
        min_count = int(sampling.get("min_count", 1) or 1)
        if mode == "按数量":
            keep_count = int(value)
        else:
            ratio = value / 100 if value > 1 else value
            keep_count = int((len(matched) * ratio) + 0.999999)
        keep_count = max(min_count, keep_count)
        matched = random.sample(matched, min(keep_count, len(matched)))
    return matched


def _apply_plan_overtime(rows, plan, headers, duration_col):
    overtime = plan.get("overtime") or {}
    if not overtime.get("enabled"):
        return rows, []
    send_col = _resolve_col(overtime.get("send_column"), headers, 1)
    process_col = _resolve_col(overtime.get("process_column"), headers, 28)
    duration_source_col = _resolve_col(overtime.get("duration_column"), headers, 0) if overtime.get("duration_column") else 0
    id_col = _resolve_col(overtime.get("id_column"), headers, 2)
    threshold = float(overtime.get("threshold_minutes", 20) or 20)
    overtime_ids = []
    updated_rows = []
    for row in rows:
        if overtime.get("mode") == "使用已有处理时长列" and duration_source_col:
            raw_minutes = _row_value(row, duration_source_col)
            try:
                minutes = round(float(raw_minutes), 2)
                row = _set_row_value(row, duration_col, minutes)
            except (TypeError, ValueError):
                minutes = None
                row = _set_row_value(row, duration_col, "时长格式错误")
        else:
            send_time = _parse_datetime(_row_value(row, send_col))
            process_time = _parse_datetime(_row_value(row, process_col))
            if send_time and process_time:
                minutes = round((process_time - send_time).total_seconds() / 60, 2)
                row = _set_row_value(row, duration_col, minutes)
            else:
                minutes = None
                row = _set_row_value(row, duration_col, "时间格式错误")
        if minutes is not None:
            row = _set_row_value(row, duration_col, minutes)
            if minutes > threshold:
                overtime_ids.append(_row_text(row, id_col))
        updated_rows.append(row)
    return updated_rows, overtime_ids


def process_file(input_path, output_dir=None, inspector="未命名质检员", run_date=None, config_path=None, progress_callback=None, scheme_id=None):
    def progress(message):
        if progress_callback:
            progress_callback(message)

    input_path = Path(input_path)
    output_dir = Path(output_dir) if output_dir else input_path.parent
    run_date = run_date or datetime.now()
    inspector = inspector.strip() or "未命名质检员"
    config = load_config(config_path)
    scheme = get_active_scheme(config, scheme_id)
    validation = validate_config_for_file(input_path, scheme_id, config_path)
    if not validation["ok"]:
        raise ValueError("配置校验未通过：\n" + "\n".join(validation["errors"]))
    fields = {**(scheme.get("fields") or DEFAULT_SCHEME["fields"]), **_fields_from_items(scheme.get("field_items"))}
    rule_values, field_overrides = _values_from_rules(scheme.get("value_rules"))
    fields.update({key: value for key, value in field_overrides.items() if value})
    values = {**(scheme.get("values") or DEFAULT_SCHEME["values"]), **rule_values}
    companies = set(values.get("company_names") or config.get("companies") or COMPANIES)
    overtime_threshold = float(scheme.get("overtime_threshold_minutes") or config.get("overtime_threshold_minutes", 20) or 20)
    invalid_sample_rate = float(scheme.get("invalid_sample_rate", 0.2) or 0.2)
    invalid_sample_min = int(scheme.get("invalid_sample_min", 1) or 1)

    progress("读取 Excel 文件...")
    wb = load_workbook(input_path, read_only=True, data_only=False)
    src_ws = wb.worksheets[0]
    if src_ws.max_row == 1 and src_ws.max_column == 1:
        src_ws.reset_dimensions()

    progress("读取并筛选数据...")
    source_rows = src_ws.iter_rows(values_only=True)
    try:
        header = tuple(next(source_rows))
    except StopIteration:
        raise ValueError("源文件没有可处理的数据")

    headers = list(header)
    max_col = max(src_ws.max_column or 0, len(header))
    total_rows = src_ws.max_row - 1 if src_ws.max_row else "未知"
    progress(f"源表数据：{total_rows} 行，{max_col} 列")
    col_keep = _resolve_col(fields.get("keep"), headers, 33)
    col_company = _resolve_col(fields.get("company"), headers, 8)
    col_send = _resolve_col(fields.get("send_time"), headers, _find_col(headers, "工单派发时间", 1))
    col_process = _resolve_col(fields.get("process_time"), headers, _find_col(headers, "客服部处理时间", 28))
    col_event_nature = _resolve_col(fields.get("event_nature"), headers, 15)
    col_r = _resolve_col(fields.get("valid_scope"), headers, _find_col(headers, "是否符合舆情范围", 18))
    col_s = _resolve_col(fields.get("marketing"), headers, _find_col(headers, "是否营销类舆情事件", 19))
    col_reminder = _resolve_col(fields.get("reminder"), headers, 20)
    col_u = _resolve_col(fields.get("duplicate"), headers, _find_col(headers, "是否为重复事件", 21))
    col_category = _resolve_col(fields.get("event_category"), headers, 25)
    col_id = _resolve_col(fields.get("id"), headers, _find_col(headers, "编号", 2))
    special_target_columns = _resolve_cols(fields.get("special_targets") or config.get("special_target_columns"), headers, [3, 11])

    progress(f"当前方案：{scheme.get('name', DEFAULT_SCHEME['name'])}")
    progress(f"专项匹配列：{', '.join(get_column_letter(col) for col in special_target_columns)}")

    filtered_rows = []
    scanned_count = 0
    ag_count = 0
    for row in source_rows:
        scanned_count += 1
        if not _row_text(row, col_keep):
            continue
        ag_count += 1
        if _row_text(row, col_company) in companies:
            filtered_rows.append(tuple(row))
    wb.close()
    progress(f"AG 列非空：{ag_count} 行；公司筛选后保留：{len(filtered_rows)} 行")

    duration_col = max_col + 1
    header = _set_row_value(header, duration_col, "处理时长(分钟)")
    processed_rows = list(filtered_rows)
    month_plan = get_monthly_plan(run_date.month, config.get("monthly_special_plans"))
    review_plans = _normalize_review_plans(scheme)
    plan_results = []
    all_overtime_ids = []

    progress("执行质检计划...")
    for plan in review_plans:
        if not plan.get("enabled", True):
            continue
        run_plan = copy_json(plan)
        if run_plan.get("role") == "monthly_special" and run_plan.get("match_type") == "按月份专项关键词":
            run_plan["name"] = month_plan["name"]
            run_plan["keywords"] = "，".join(month_plan.get("keywords") or [])
            run_plan["keyword_columns"] = ",".join(get_column_letter(col) for col in special_target_columns)
        matched_rows = _run_review_plan(processed_rows, run_plan, headers, month_plan if run_plan.get("match_type") == "按月份专项关键词" else None)
        matched_rows, overtime_ids = _apply_plan_overtime(matched_rows, run_plan, headers, duration_col)
        all_overtime_ids.extend([item for item in overtime_ids if item])
        plan_results.append({"plan": run_plan, "rows": matched_rows, "overtime_ids": overtime_ids})
        progress(f"{run_plan.get('name', '质检计划')}：命中 {len(matched_rows)} 行")

    progress("生成日报送文本...")
    a = sum(1 for row in processed_rows if _row_text(row, 1))
    b = _count(processed_rows, col_r, values.get("valid_no", "否"))
    c = _count(processed_rows, col_event_nature, values.get("positive_event", "正面事件"))
    d = _countifs(processed_rows, [(col_event_nature, values.get("negative_event", "负面事件")), (col_s, values.get("marketing_no", "否"))])
    e = _countifs(processed_rows, [(col_s, values.get("marketing_yes", "是")), (col_r, values.get("valid_yes", "是"))])
    f = _countifs(processed_rows, [(col_s, values.get("marketing_yes", "是")), (col_r, values.get("valid_yes", "是")), (col_u, values.get("duplicate_yes", "是"))])
    g = _count(processed_rows, col_reminder, values.get("reminder", "舆情提醒"))
    h = _countifs(processed_rows, [(col_category, values.get("livelihood_event", "民生类舆情")), (col_r, values.get("valid_yes", "是")), (col_s, values.get("marketing_yes", "是"))])

    report_text = (
        f"{run_date.year}年{run_date.month}月{run_date.day}日，南方分中心共收到舆情工单待办{a}件，"
        f"其中无效事件{b}件、正面事件{c}件、负面非营销类事件{d}件，舆情提醒{g}件，"
        f"负面营销类舆情{e}件（重复事件{f}件，民生事件{h}件）。"
    )

    out_wb = Workbook()
    out_wb.remove(out_wb.active)
    progress("导出最终质检明细...")
    for result in plan_results:
        plan = result["plan"]
        if not plan.get("output_sheet", True):
            continue
        sheet_name = plan.get("name", "质检计划")
        if plan.get("role") == "monthly_special" and not sheet_name.endswith("复核"):
            sheet_name += "复核"
        _append_review_sheet(
            out_wb,
            _safe_sheet_name(sheet_name),
            header,
            result["rows"],
            inspector,
            plan=plan,
            source_headers=headers,
            duration_col=duration_col,
            month_plan=month_plan if plan.get("match_type") == "按月份专项关键词" else None,
        )
    if not out_wb.worksheets:
        _append_review_sheet(out_wb, "质检结果", header, [], inspector, source_headers=headers, duration_col=duration_col)

    yesterday = run_date - timedelta(days=1)
    file_name = f"{yesterday.year}年{yesterday.month}月{yesterday.day}日8点-{run_date.year}年{run_date.month}月{run_date.day}日8点舆情质检明细({inspector}).xlsx"
    output_path = output_dir / file_name
    out_wb.save(output_path)
    progress("保存完成")

    return {
        "output_path": str(output_path),
        "report_text": report_text,
        "special_name": month_plan["name"],
        "special_count": next((len(item["rows"]) for item in plan_results if item["plan"].get("role") == "monthly_special"), 0),
        "reminder_count": next((len(item["rows"]) for item in plan_results if item["plan"].get("role") == "reminder_review"), 0),
        "invalid_count": next((len(item["rows"]) for item in plan_results if item["plan"].get("role") == "invalid_review"), 0),
        "overtime_count": len(all_overtime_ids),
        "overtime_ids": all_overtime_ids,
        "plan_results": [
            {"name": item["plan"].get("name", "质检计划"), "count": len(item["rows"]), "overtime_count": len(item["overtime_ids"])}
            for item in plan_results
        ],
    }
